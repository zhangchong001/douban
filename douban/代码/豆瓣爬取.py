#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib.request
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver

#模拟登录加验证

# 导入鼠标事件类
from selenium.webdriver import ActionChains
import time

# 加速度函数
def get_tracks(distance):
    """
    拿到移动轨迹，模仿人的滑动行为，先匀加速后匀减速
    匀变速运动基本公式：
    ①v=v0+at
    ②s=v0t+½at²
    """
    # 初速度
    v = 0
    # 单位时间为0.3s来统计轨迹，轨迹即0.3内的位移
    t = 0.3
    # 位置/轨迹列表,列表内的一个元素代表0.3s的位移
    tracks = []
    # 当前的位移
    current = 0
    # 到达mid值开始减速
    mid = distance*4/5
    while current < distance:
        if current < mid:
            # 加速度越小,单位时间内的位移越小,模拟的轨迹就越多越详细
            a = 2
        else:
            a = -3

        # 初速度
        v0 = v
        # 0.3秒内的位移
        s = v0*t+0.5*a*(t**2)
        # 当前的位置
        current += s
        # 添加到轨迹列表
        tracks.append(round(s))
        # 速度已经达到v，该速度作为下次的初速度
        v = v0 + a*t
    return tracks
    # tracks: [第一个0.3秒的移动距离,第二个0.3秒的移动距离,...]


# 1、打开豆瓣官网 - 并将窗口最大化
browser = webdriver.Chrome()
browser.maximize_window()
browser.get('https://www.douban.com/')

# 2、切换到iframe子页面
login_frame = browser.find_element_by_xpath('//*[@id="anony-reg-new"]/div/div[1]/iframe')
browser.switch_to.frame(login_frame)

# 3、密码登录 + 用户名 + 密码 + 登录豆瓣
browser.find_element_by_xpath('/html/body/div[1]/div[1]/ul[1]/li[2]').click()
browser.find_element_by_xpath('//*[@id="username"]').send_keys('13148144907')
browser.find_element_by_xpath('//*[@id="password"]').send_keys('zc@5201314')
browser.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[5]/a').click()
time.sleep(4)

# 4、切换到新的iframe子页面 - 滑块验证
auth_frame = browser.find_element_by_xpath('//*[@id="TCaptcha"]/iframe')
browser.switch_to.frame(auth_frame)
# 5、按住开始滑动位置按钮 - 先移动180个像素
element = browser.find_element_by_xpath('//*[@id="tcaptcha_drag_button"]')
ActionChains(browser).click_and_hold(on_element=element).perform()
#move_to_element_with_offset():移动到距离某个元素左上角多少距离的位置
ActionChains(browser).move_to_element_with_offset(to_element=element,xoffset=180,yoffset=0).perform()
# 6、使用加速度函数移动剩下的距离
tracks = get_tracks(25)
for track in tracks:
    #开始移动 move_by_offset() 鼠标从当前位置移动多少距离
    ActionChains(browser).move_by_offset(xoffset=track,yoffset=0).perform()
# 7、延迟释放鼠标: release()
time.sleep(0.5)
ActionChains(browser).release().perform()

#数据爬取

#Some User Agents
hds=[{'User-Agent': 'mozilla/5.0 (windows nt 6.1; wow64) applewebkit/537.36 (khtml, like gecko) chrome/27.0.1453.94 safari/537.36'}, \
       {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}, \
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num=0;
    book_list=[]
    try_times=0
    
    while(page_num < 5):
        url='http://www.douban.com/tag/'+urllib.request.quote(book_tag)+'/book?start='+str(page_num*15)
        time.sleep(np.random.rand()*5)
        
        try:
            req = urllib.request.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib.request.urlopen(req).read().decode('utf-8')
            plain_text=str(source_code)   
        except (urllib.request.HTTPError, urllib.request.URLError) as e:
            print(e)
            continue
        
        soup = BeautifulSoup(plain_text, 'lxml')
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href')
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num ='0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print('正在下载第 %d 页信息' % page_num)
    return book_list


def get_people_num(url):
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib.request.urlopen(req).read()
        plain_text=str(source_code)   
    except (urllib.request.HTTPError, urllib.request.URLError) as e:
        print(e)
    soup = BeautifulSoup(plain_text,'lxml')
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)
        print('%s类图书完成爬取' %(book_tag))
        time.sleep(60)
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    wb = Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i])) #utf8->unicode
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)




if __name__=='__main__':
    book_tag_lists = ['大数据','C语言','python','web','java']
    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)
    print("程序结束，文件爬取完毕")